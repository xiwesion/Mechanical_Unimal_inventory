"""
Template Handler Module
Handles Excel template generation and import/export
"""
from pathlib import Path
from datetime import datetime
import json
from io import BytesIO
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class TemplateHandler:
    """Handles Excel template and data import/export"""
    
    TEMPLATE_DIR = Path(__file__).parent.parent / "templates"
    TEMPLATE_FILE = TEMPLATE_DIR / "equipment_template.xlsx"
    
    def __init__(self):
        """Initialize template handler"""
        self._ensure_template()
    
    def _ensure_dir(self):
        """Ensure templates directory exists"""
        self.TEMPLATE_DIR.mkdir(exist_ok=True)
    
    def _ensure_template(self):
        """Ensure default template exists"""
        self._ensure_dir()
        if not self.TEMPLATE_FILE.exists():
            self._create_default_template()
    
    def _create_default_template(self):
        """Create default Excel template"""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Equipment"
            
            # Define styles
            header_fill = PatternFill(start_color="A5D6A7", end_color="A5D6A7", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            # Define columns
            columns = [
                ("Nama Equipment*", 40),
                ("Jumlah/Quantity*", 15),
                ("Merk", 20),
                ("Type", 15),
                ("BoM", 15),
                ("Harga Satuan*", 15),
                ("Harga Keseluruhan", 18),
                ("Kategori", 20),
                ("Keterangan", 30),
                ("Status", 12)
            ]
            
            # Write headers
            for col_num, (col_name, col_width) in enumerate(columns, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = col_name
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment
                cell.border = border
                ws.column_dimensions[get_column_letter(col_num)].width = col_width
            
            # Add instruction row
            ws.row_dimensions[2].height = 30
            instruction = "Isikan data equipment di bawah ini. Kolom dengan (*) wajib diisi."
            ws.merge_cells('A2:J2')
            inst_cell = ws['A2']
            inst_cell.value = instruction
            inst_cell.font = Font(italic=True, size=10, color="666666")
            inst_cell.alignment = Alignment(wrap_text=True)
            
            # Add example data (optional)
            example_data = [
                ["Mesin Bubut", 2, "TOKOTA", "CNC", "M01", 50000000, 100000000, "Mesin", "Kondisi baik", "Aktif"],
                ["Alat Ukur Digital", 5, "Mitutoyo", "Caliper", "M02", 250000, 1250000, "Alat Ukur", "", "Aktif"],
            ]
            
            for row_num, row_data in enumerate(example_data, 3):
                for col_num, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.value = value
                    cell.border = border
                    if col_num in [2, 6, 7]:  # Quantity and price columns
                        cell.number_format = '#,##0'
                    cell.alignment = Alignment(wrap_text=True)
            
            # Add data validation (optional)
            ws.row_dimensions[1].height = 25
            
            # Save template
            wb.save(self.TEMPLATE_FILE)
            print(f"Template created: {self.TEMPLATE_FILE}")
        except Exception as e:
            print(f"Error creating template: {e}")
    
    def get_template(self) -> bytes:
        """Get template as bytes for download"""
        try:
            with open(self.TEMPLATE_FILE, 'rb') as f:
                return f.read()
        except Exception as e:
            print(f"Error reading template: {e}")
            return None
    
    def import_equipment(self, file, lab_id: str, equipment_manager=None) -> tuple[bool, str, list]:
        """
        Import equipment from Excel file
        
        Args:
            file: Uploaded file object
            lab_id: Target lab ID
            equipment_manager: EquipmentManager instance (optional, will create new if not provided)
        
        Returns:
            (success, message, imported_equipment_list)
        """
        try:
            # Read Excel file
            df = pd.read_excel(file, sheet_name="Equipment")
            
            # Validate and clean data
            df = df.dropna(subset=['Nama Equipment*'])  # Required field
            df = df.fillna('')
            
            imported = []
            errors = []
            
            # Use provided equipment_manager or create new one
            if equipment_manager is None:
                from modules.equipment_manager import EquipmentManager
                em = EquipmentManager()
            else:
                em = equipment_manager
            
            for idx, row in df.iterrows():
                try:
                    # Prepare equipment data
                    equipment_data = {
                        'nama': str(row.get('Nama Equipment*', '')).strip(),
                        'jumlah': float(row.get('Jumlah/Quantity*', 0)) or 0,
                        'merk': str(row.get('Merk', '')).strip(),
                        'type': str(row.get('Type', '')).strip(),
                        'bom': str(row.get('BoM', '')).strip(),
                        'harga_satuan': float(row.get('Harga Satuan*', 0)) or 0,
                        'kategori': str(row.get('Kategori', 'Lainnya')).strip(),
                        'keterangan': str(row.get('Keterangan', '')).strip(),
                        'status': str(row.get('Status', 'active')).lower().replace(' ', '_')
                    }
                    
                    # Validate
                    if not equipment_data['nama']:
                        errors.append(f"Row {idx + 3}: Nama Equipment harus diisi")
                        continue
                    
                    if equipment_data['jumlah'] < 0:
                        errors.append(f"Row {idx + 3}: Jumlah tidak boleh negatif")
                        continue
                    
                    if equipment_data['harga_satuan'] < 0:
                        errors.append(f"Row {idx + 3}: Harga satuan tidak boleh negatif")
                        continue
                    
                    # Calculate total
                    equipment_data['harga_keseluruhan'] = equipment_data['jumlah'] * equipment_data['harga_satuan']
                    
                    # Add equipment
                    eq_id = em.add_equipment(lab_id, equipment_data)
                    
                    if eq_id:
                        imported.append({
                            'equipment_id': eq_id,
                            'nama': equipment_data['nama'],
                            'jumlah': equipment_data['jumlah']
                        })
                    else:
                        errors.append(f"Row {idx + 3}: Gagal menambahkan equipment")
                
                except Exception as e:
                    errors.append(f"Row {idx + 3}: {str(e)}")
            
            # Prepare message
            message = f"Berhasil import {len(imported)} equipment"
            if errors:
                message += f" dengan {len(errors)} error:\n" + "\n".join(errors[:5])
                if len(errors) > 5:
                    message += f"\n... dan {len(errors) - 5} error lainnya"
            
            return len(imported) > 0, message, imported
            
        except Exception as e:
            return False, f"Error: {str(e)}", []
    
    def export_equipment(self, equipment_list: list, include_consumption: bool = False) -> bytes:
        """
        Export equipment data to Excel
        
        Args:
            equipment_list: List of equipment dictionaries
            include_consumption: Include consumption data
        
        Returns:
            Excel file as bytes
        """
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Equipment"
            
            # Define styles
            header_fill = PatternFill(start_color="A5D6A7", end_color="A5D6A7", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Define columns
            if include_consumption:
                columns = [
                    "Equipment ID", "Nama", "Jumlah", "Merk", "Type", "BoM",
                    "Harga Satuan", "Harga Total", "Kategori", "Status", 
                    "Dibuat Tanggal", "Terakhir Diubah", "Keterangan"
                ]
            else:
                columns = [
                    "Equipment ID", "Nama", "Jumlah", "Merk", "Type", "BoM",
                    "Harga Satuan", "Harga Total", "Kategori", "Status", "Keterangan"
                ]
            
            # Write headers
            for col_num, col_name in enumerate(columns, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = col_name
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                ws.column_dimensions[chr(64 + col_num)].width = 15
            
            # Write data
            for row_num, eq in enumerate(equipment_list, 2):
                ws.cell(row=row_num, column=1).value = eq.get('equipment_id', '')
                ws.cell(row=row_num, column=2).value = eq.get('nama', '')
                ws.cell(row=row_num, column=3).value = float(eq.get('jumlah', 0))
                ws.cell(row=row_num, column=4).value = eq.get('merk', '')
                ws.cell(row=row_num, column=5).value = eq.get('type', '')
                ws.cell(row=row_num, column=6).value = eq.get('bom', '')
                ws.cell(row=row_num, column=7).value = float(eq.get('harga_satuan', 0))
                ws.cell(row=row_num, column=8).value = float(eq.get('harga_keseluruhan', 0))
                ws.cell(row=row_num, column=9).value = eq.get('kategori', '')
                ws.cell(row=row_num, column=10).value = eq.get('status', '')
                
                if include_consumption:
                    ws.cell(row=row_num, column=11).value = eq.get('created_date', '')
                    ws.cell(row=row_num, column=12).value = eq.get('last_modified', '')
                    ws.cell(row=row_num, column=13).value = eq.get('keterangan', '')
                else:
                    ws.cell(row=row_num, column=11).value = eq.get('keterangan', '')
            
            # Save to bytes
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            return output.getvalue()
            
        except Exception as e:
            print(f"Error exporting equipment: {e}")
            return None
    
    def export_consumption_report(self, consumption_data: dict, lab_name: str = "") -> bytes:
        """
        Export consumption/depletion report
        
        Args:
            consumption_data: Consumption data dictionary
            lab_name: Lab name for report
        
        Returns:
            Excel file as bytes
        """
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Consumption Report"
            
            # Add title
            ws.merge_cells('A1:F1')
            title_cell = ws['A1']
            title_cell.value = f"Laporan Konsumsi Equipment - {lab_name}"
            title_cell.font = Font(bold=True, size=14, color="FFFFFF")
            title_cell.fill = PatternFill(start_color="66BB6A", end_color="66BB6A", fill_type="solid")
            
            # Add date
            ws.merge_cells('A2:F2')
            date_cell = ws['A2']
            date_cell.value = f"Tanggal: {datetime.now().strftime('%d %B %Y')}"
            date_cell.font = Font(italic=True)
            
            # Headers
            headers = ["Equipment ID", "Nama", "Lab", "Total Dikonsumsi", "Terakhir Dikonsumsi", "Jumlah Movements"]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col_num)
                cell.value = header
                cell.fill = PatternFill(start_color="A5D6A7", end_color="A5D6A7", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
            
            # Data
            for row_num, (eq_id, data) in enumerate(consumption_data.items(), 5):
                ws.cell(row=row_num, column=1).value = eq_id
                ws.cell(row=row_num, column=2).value = data.get('equipment_name', '')
                ws.cell(row=row_num, column=3).value = data.get('lab_id', '')
                ws.cell(row=row_num, column=4).value = data.get('total_consumed', 0)
                ws.cell(row=row_num, column=5).value = data.get('last_consumption', '')
                ws.cell(row=row_num, column=6).value = len(data.get('movements', []))
            
            # Save to bytes
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            return output.getvalue()
            
        except Exception as e:
            print(f"Error exporting consumption report: {e}")
            return None
